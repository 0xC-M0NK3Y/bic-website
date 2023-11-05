import openpyxl
from openpyxl_image_loader import SheetImageLoader
import sys
import os
import cv2

MAX_HEIGHT = 350
MAX_WIDTH = 55

def fix_image(f):
	img = cv2.imread(f, cv2.IMREAD_UNCHANGED)
	img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
	h, w, _ = img.shape
	coef = min(MAX_HEIGHT/h, MAX_WIDTH/w)
	img2 = cv2.resize(img, (int(w*coef), int(h*coef)), interpolation=cv2.INTER_AREA)
	tmp = f[:-3] + "jpg"
	cv2.imwrite(tmp, img2, [int(cv2.IMWRITE_JPEG_QUALITY), 100])
	os.remove(f)

def main():

	if len(sys.argv) != 3:
		print(f'Usage: {sys.argv[0]} <bic.xlsx> <filename for out csv>')
		exit()

	os.mkdir('images')
	try:
		pxl_doc = openpyxl.load_workbook(sys.argv[1], data_only=True)
	except:
		print('Bad .xlsx document')
		exit()

	sheet = pxl_doc['achats']
	image_loader = SheetImageLoader(sheet)

	out = open(sys.argv[2], 'w')

	for row in range(2, sheet.max_row):
		print(f"Ligne : {row+1} : {sheet[row+1][2].value}", flush=True)
		# detect end with empty colors cell
		if sheet[row+1][4].value == None:
			print(f'Fin du fichier excel détectée ligne {row+1}.', file=sys.stderr)
			break
		# save image
		try:
			image = image_loader.get(f'D{row+1}')
			path = f'images/bic_{sheet[row+1][19].value}.png'
			image.save(path)
			fix_image(path)
			path = path[:-3] + 'jpg'
		except:
			print(f'Erreur avec l\'image ligne {row} (la ligne sera pas prise en compte)', file=sys.stderr)
			continue

		# print csv line
		for col in sheet.iter_cols(1, sheet.max_column):
			if col[0].value == None:
				continue
			if col[0].value == 'image':
				print(path, end='|', file=out)
			else:
				if col[0].value == 'latitude' and col[row].value == None:
					continue
				if col[0].value == 'longitude' and col[row].value == None:
					continue
				if type(col[row].value) == str:
					print(col[row].value.strip(), end='|', file=out)
				else:
					print(col[row].value, end='|', file=out)
		print('', file=out)


if __name__ == '__main__':
	main()
